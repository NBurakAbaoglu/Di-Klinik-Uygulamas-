from tkinter import *
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
from Crypto.Random import get_random_bytes
import base64
import openpyxl
from datetime import datetime
import re
import ctypes
import subprocess
import os
import sys
import traceback
class Uygulama:
    def __init__(self):
        self.key = b'1234567890abcdef'
        self.iv = b'abcdef1234567890'
        self.bugün=datetime.today()
        self.gun=self.bugün.day
        self.ay=self.bugün.month
        self.ekranbir()
        self.ayinbiri()
        self.uyarı = None
        self.checkbuttons = []
        self.durum=True
        #self.tk_tutari=0
    def get_resource_path(self,relative_path):
        try:
            base_path = sys._MEIPASS
        except AttributeError:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)
        
        
       
        
        
    

    def ekranbir(self):
        self.ekran1 = Tk()
        self.ekran1.title("Klinik Adı")
        self.ekran1.configure(bg="black")
        self.ekrankapama = 0
        try:
            self.logo_yolu=self.get_resource_path(r"Klinik giriş ekranı resmi")
            self.logo = PhotoImage(file=self.logo_yolu)
        except Exception as e:
            with open("hata_log.txt", "a") as f:
                f.write(f"Error loading png: {e}\n")
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("V.1")
        self.label = Label(self.ekran1, image=self.logo, bd=0)
        self.label.place(width=800, height=600)
        self.label2 = Label(text="V1.7", fg="white", font='Arial', background="black")
        self.label2.place(relx=1.0, rely=1.0, anchor="se")
        self.ekran1.geometry("900x900+500+50")
        # Dinamik dosya yolunu al
        

        self.icon_yolu = self.get_resource_path("Klinik Logo resmi")
        
        try:
            # İkonu yükle
            self.ekran1.iconbitmap(self.icon_yolu)
        except Exception as e:
            with open("hata_log.txt", "a") as f:
                f.write(f"Error loading icon: {e}\n")

        self.ekran1.resizable(height=False, width=False)
        self.button = Button(text="GİRİŞ YAP", font=('italic', 50), command=self.ekraniki, background="gold", 
                             activebackground="white", activeforeground="black", bd=5, highlightbackground="white")
        self.button.place(x=200, y=600)
        self.ekran1.mainloop()

    def ekraniki(self):
        self.ekrankapama += 1
        if self.ekrankapama == 1:
            self.ekran1.destroy()
        self.ekran2 = Tk()
        self.ekran2.title("Klinik Adı")
        self.ekran2.configure(bg="#666666")
        self.ekran2.geometry("800x400+600+250")

        self.icon_yolu = self.get_resource_path("Klinik Logo resmi")
        
        try:
            # İkonu yükle
            self.ekran1.iconbitmap(self.icon_yolu)
        except Exception as e:
            print(f"Error loading icon: {e}")

        self.grs = Label(text="KULLANICI ADI: ", fg="white", font=("Arial", 20), background="#666666", justify="center")
        self.sifre = Label(text="KULLANICI ŞİFRE: ", fg="white", font=("Arial", 20), background="#666666")


        self.label3 = Label(text="V1.7", fg="white", font='Arial', background="#666666")
        self.label3.pack(expand=1,anchor='se')

        self.grs.place(x=50, y=50)
        self.sifre.place(x=30, y=150)

        self.grs_veri = Entry(font=("Arial", 20))
        self.grs_veri.place(x=280, y=50)

        self.sifre_veri = Entry(font=("Arial", 20))
        self.sifre_veri.place(x=280, y=150)

        self.grs_buton = Button(text="GİRİŞ YAP", font=("Arial", 20), command=self.giris)
        self.grs_buton.place(x=50, y=250)

        self.kytol = Button(text="ŞİFREMİ UNUTTUM", font=("Arial", 20), command=self.kayıt)
        self.kytol.place(x=250, y=250)
        
        self.temizle = Button(text="TEMİZLE", font=("Arial", 20),command=self.temizle)
        self.temizle.place(x=550, y=250)

        self.ekran2.mainloop()
    def temizle(self):
        self.grs_veri.delete(0, END)
        self.sifre_veri.delete(0, END)
    def giris(self):
        dosya=self.get_resource_path(r'Kliniktkinter\encrypted_message.txt')
        with open(dosya, 'r', encoding='utf-8') as file:
            encrypted_name = file.readline().strip()  
            encrypted_password = file.readline().strip()  
            self.ok_name = self.decrypt_message(encrypted_name)  
            self.ok_password = self.decrypt_message(encrypted_password)  

        
        if self.grs_veri.get() == self.ok_name and self.sifre_veri.get() == self.ok_password:
            self.islemler()
        else:
            messagebox.showinfo("BİLDİRİM", "KULLANICI ADI VEYA ŞİFRE YANLIŞ VEYA EKSİK GİRİLMİŞTİR")


    def kayıt(self):
        self.sicilno = 443820
        self.ekran2.destroy()
        self.ekran3 = Tk()
        self.ekran3.title("Klinik Adı")
        self.ekran3.configure(bg="#666666")
        self.ekran3.geometry("800x400+600+250")
        self.icon_yolu = self.get_resource_path("Klinik Logo resmi")
        
        try:
            # İkonu yükle
            self.ekran1.iconbitmap(self.icon_yolu)
        except Exception as e:
            print(f"Error loading icon: {e}")

        self.grs = Label(text="YENİ KULLANICI İSİM: ", fg="white", font=("Arial", 20), background="#666666", justify="center")
        self.sifre = Label(text="YENİ KULLANICI ŞİFRE: ", fg="white", font=("Arial", 20), background="#666666")
        self.sicil_no = Label(text="SİCİL NO GİRİNİZ: ", fg="white", font=("Arial", 20), background="#666666")

        self.grs.place(x=30, y=50)
        self.sifre.place(x=30, y=150)
        self.sicil_no.place(x=30, y=250)

        self.grs_veri = Entry(font=("Arial", 20))
        self.grs_veri.place(x=350, y=50)

        self.label3 = Label(text="V1.7", fg="white", font='Arial', background="#666666")
        self.label3.pack(expand=1,anchor='se')

        self.sifre_veri = Entry(font=("Arial", 20))
        self.sifre_veri.place(x=350, y=150)
        
        self.sicil_veri = Entry(font=("Arial", 20))
        self.sicil_veri.place(x=350, y=250)

        self.kyt_buton = Button(text="KAYIT OL", font=("Arial", 20), command=self.sorgu_kyt)
        self.kyt_buton.place(x=150, y=300)
        self.geri_don_sy2=Button(text="GERİ DÖN", font=("Arial", 20), command=self.geri_don)
        self.geri_don_sy2.place(x=450, y=300)
        self.ekran3.mainloop()
    def geri_don(self):
        self.ekran3.destroy()
        self.ekraniki()
    def sorgu_kyt(self):
        try:
            if int(self.sicil_veri.get()) == self.sicilno:
                self.name = self.encrypt_message(self.grs_veri.get())
                self.password =self.encrypt_message( self.sifre_veri.get())
                dosya=self.get_resource_path(r"Kliniktkinter\encrypted_message.txt")
                with open(dosya, "w", encoding="utf-8") as file:
                    file.write(f"{self.name}\n{self.password}")
                messagebox.showinfo("BİLDİRİM", "KAYIT BAŞARIYLA TAMAMLANMIŞTIR")
                self.ekran3.destroy()
                self.ekraniki()
            else:
                messagebox.showinfo("BİLDİRİM","SİCİL NUMARASI HATALI EKSİK VEYA HATALI TUŞLADINIZ")
        except ValueError:
            messagebox.showinfo("BİLDİRİM","GEÇERSİZ GİRİŞ! SİCİL NO YALNIZCA HARFLERDEN OLUŞMALIDIR")
        

    def encrypt_message(self, message):
        
        cipher = AES.new(self.key, AES.MODE_CBC, self.iv)
        padded_message = pad(message.encode('utf-8'), AES.block_size)
        encrypted_bytes = cipher.encrypt(padded_message)
        encrypted_message = base64.b64encode(encrypted_bytes).decode('utf-8')
        return encrypted_message
    
    def decrypt_message(self, encrypted_message):
        cipher = AES.new(self.key, AES.MODE_CBC, self.iv)
        encrypted_bytes = base64.b64decode(encrypted_message)
        decrypted_message = unpad(cipher.decrypt(encrypted_bytes), AES.block_size).decode('utf-8')
        return decrypted_message
    def ayinbiri(self):
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)

        self.sayfa_gider=self.excel_dosya["GİDER"]
        self.satir_sayisi=2267
        self.gider_satir_sayisi=1
        self.toplam_ay_gelir=0
        self.sayfa=self.excel_dosya["GELİR"]
        self.brüt_sayfa=self.excel_dosya["NET GELİR"]


        self.toplam_islem_ücreti=0
        self.toplam_ödenen=0
        self.aylik_net_gelir=0
        self.aylik_maliyet=0
        self.doktor1_prim=0
        self.doktor2_prim=0
        self.doktor3_prim=0
        self.doktor4_prim=0
        self.doktor5_prim=0
        self.gdr_satır_sayisi_aylik=1
        self.aylar = [
    "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
        baslangıc_sutün=None
        bitis_sutün=None
        baslangıc_sütün_gdr=None
        bitis_sutün_gdr=None
        self.brüt_satır=1
        while True:
          
            self.veri_varmi = any(self.sayfa.cell(row=self.satir_sayisi, column=col).value is not None for col in range(1, 16))
            
            for col in range(1, 16): 
                self.ay_deger = self.sayfa.cell(row=self.satir_sayisi, column=col).value
                if self.ay_deger in self.aylar:  
                    baslangıc_sutün = self.satir_sayisi+1
            
            if not self.veri_varmi:
                bitis_sutün = self.satir_sayisi
                break
            
            self.satir_sayisi += 1  
        while True:
          
                self.veri_varmi = any(self.sayfa_gider.cell(row=self.gdr_satır_sayisi_aylik, column=col).value is not None for col in range(1, 11))
                
                for col in range(1, 16): 
                    self.ay_deger = self.sayfa_gider.cell(row=self.gdr_satır_sayisi_aylik, column=col).value
                    if self.ay_deger in self.aylar:  
                        baslangıc_sütün_gdr = self.gdr_satır_sayisi_aylik+1
                
                if not self.veri_varmi:
                    bitis_sutün_gdr = self.gdr_satır_sayisi_aylik
                    break
                
                self.gdr_satır_sayisi_aylik += 1  
        while True:
          
                self.veri_varmi = any(self.brüt_sayfa.cell(row=self.brüt_satır, column=col).value is not None for col in range(1, 16))
                
                if not self.veri_varmi:
                    break
                
                self.brüt_satır += 1  
        if self.ay==1:
            for col in range(1, 18):
                self.sayfa.cell(self.satir_sayisi, col, value=self.bugün.year)
                self.sayfa_gider.cell(self.gdr_satır_sayisi_aylik, col, value=self.bugün.year)
                self.brüt_sayfa.cell(self.brüt_satır, col, value=self.bugün.year)
            self.satir_sayisi += 1 
            self.gdr_satır_sayisi_aylik += 1 
            self.brüt_satır += 1 
        if self.gun == 1:
            for col in range(1, 16):
                self.sayfa.cell(self.satir_sayisi, col, value=self.aylar[self.ay - 1])
                self.sayfa_gider.cell(self.gdr_satır_sayisi_aylik, col, value=self.aylar[self.ay - 1])
            self.satir_sayisi += 1 
            self.gdr_satır_sayisi_aylik += 1 
            
            benzersiz_kayitlar = set()
            # Sonuç listesi
            self.dizi = []

            # Başlangıç sütunundan bitiş sütununa kadar döngü
            i = baslangıc_sutün
            while i < bitis_sutün:
                # Geçici dizi oluştur
                gecici_dizi = []
                
                # İsim, işlem, gider, işlem tutarı ve ödenen değerlerini al
                isim1 = self.sayfa.cell(row=i, column=3).value
                islem1 = self.sayfa.cell(row=i, column=4).value
                gider1 = self.sayfa.cell(row=i, column=5).value
                islem_tutar1 = self.sayfa.cell(row=i, column=12).value
                odenen1 = self.sayfa.cell(row=i, column=13).value
                
                # Eğer islem_tutar1 None değilse, işlemi yap
                if islem_tutar1 is not None:
                    islem_tutar1 = int(islem_tutar1.replace("₺", "").replace(".", ""))
                else:
                    islem_tutar1 = 0  # Varsayılan değer
                
                # Eğer gider1 None değilse, integer'a çevir
                if gider1 is not None:
                    gider1 = int(gider1)
                else:
                    gider1 = 0  # Varsayılan değer
                
                # Eğer ödenen1 None değilse, integer'a çevir
                if odenen1 is not None:
                    odenen1 = int(odenen1.replace("₺", "").replace(".", ""))
                else:
                    odenen1 = 0  # Varsayılan değer
                
                # Benzersiz kayıt kontrolü
                kayit = (isim1, islem1, islem_tutar1, gider1)
                
                if kayit not in benzersiz_kayitlar:
                    benzersiz_kayitlar.add(kayit)
                    gecici_dizi.extend([isim1, islem1, islem_tutar1, gider1, odenen1])
                    self.dizi.append(gecici_dizi)
                else:
                    # Eğer kayıt zaten varsa, ödenen değerini karşılaştır ve en düşük olanı seç
                    for idx, item in enumerate(self.dizi):
                        if (item[0] == isim1 and item[1] == islem1 and item[2] == islem_tutar1 and item[3] == gider1):
                            if odenen1 < item[4]:  # Ödenen değeri karşılaştır
                                self.dizi[idx][4] = odenen1  # En düşük ödenen değerini güncelle
                
                # Bir sonraki satıra geç
                i += 1
                      
               
                
            for kayıt in self.dizi:
                deger=kayıt[2]
                deger = int(deger)
                self.toplam_islem_ücreti+=deger
                
            for kayıt in self.dizi:
                    deger=kayıt[4]
                    deger = int(deger)
                    self.toplam_ödenen+=deger

            for satir in range(baslangıc_sutün,bitis_sutün + 1): 
                    deger = self.sayfa.cell(row=satir, column=15).value  
                    if deger in self.aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.aylik_net_gelir += temiz_deger 
                    else:
                        if deger is not None:
                            self.aylik_net_gelir += int(deger)  
            for kayıt in self.dizi:
                    deger=kayıt[3]
                    deger = int(deger)
                    self.aylik_maliyet+=deger
            for satir in range(baslangıc_sutün,bitis_sutün + 1): 
                    deger = self.sayfa.cell(row=satir, column=7).value  
                    if deger in self.aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor1_prim += temiz_deger 
                    else:
                        if deger is not None:
                            self.doktor1_prim += int(deger) 

            for satir in range(baslangıc_sutün,bitis_sutün + 1): 
                    deger = self.sayfa.cell(row=satir, column=8).value  
                    if deger in self.aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor2_prim += temiz_deger 
                    else:
                        if deger is not None:
                            self.doktor2_prim += int(deger)  

            for satir in range(baslangıc_sutün,bitis_sutün + 1): 
                    deger = self.sayfa.cell(row=satir, column=9).value  
                    if deger in self.aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor3_prim += temiz_deger 
                    else:
                        if deger is not None:
                            self.doktor3_prim += int(deger)  

            for satir in range(baslangıc_sutün,bitis_sutün + 1): 
                    deger = self.sayfa.cell(row=satir, column=10).value  
                    if deger in self.aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor4_prim += temiz_deger 
                    else:
                        if deger is not None:
                            self.doktor4_prim += int(deger)  

            for satir in range(baslangıc_sutün,bitis_sutün + 1): 
                    deger = self.sayfa.cell(row=satir, column=11).value  
                    if deger in self.aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor5_prim += temiz_deger
                    else:
                        if deger is not None:
                            self.doktor5_prim += int(deger) 
                        
            self.toplam_ay_islem_noktali=f"{self.toplam_islem_ücreti:,}".replace(",", ".")
            self.toplam_ödenen_noktali=f"{self.toplam_ödenen:,}".replace(",", ".")
            self.aylik_net_gelir_noktali=f"{self.aylik_net_gelir:,}".replace(",", ".")
            self.aylik_maliyet_noktali=f"{self.aylik_maliyet:,}".replace(",", ".")
            self.doktor1_prim_noktali=f"{self.doktor1_prim:,}".replace(",", ".")
            self.doktor2_prim_noktali=f"{self.doktor2_prim:,}".replace(",", ".")
            self.doktor3_prim_noktali=f"{self.doktor3_prim:,}".replace(",", ".")
            self.doktor4_prim_noktali=f"{self.doktor4_prim:,}".replace(",", ".")
            self.doktor5_prim_noktali=f"{self.doktor5_prim:,}".replace(",", ".")
            
            
           

            self.brüt_sayfa.cell(self.brüt_satır,column=1, value=self.aylar[self.ay - 1])
            self.brüt_sayfa.cell(self.brüt_satır,column=2, value=self.toplam_ay_islem_noktali)
            self.brüt_sayfa.cell(self.brüt_satır,column=3, value=self.toplam_ödenen_noktali)
            self.brüt_sayfa.cell(self.brüt_satır,column=5, value=self.aylik_net_gelir_noktali)
            self.brüt_sayfa.cell(self.brüt_satır,column=4, value=self.aylik_maliyet_noktali)
            self.brüt_sayfa.cell(self.brüt_satır,column=6, value=self.doktor1_prim_noktali)
            self.brüt_sayfa.cell(self.brüt_satır,column=7, value=self.doktor2_prim_noktali)
            self.brüt_sayfa.cell(self.brüt_satır,column=8, value=self.doktor3_prim_noktali)
            self.brüt_sayfa.cell(self.brüt_satır,column=9, value=self.doktor4_prim_noktali)
            self.brüt_sayfa.cell(self.brüt_satır,column=10, value=self.doktor5_prim_noktali)

        
        self.excel_dosya.save(self.excel_dosya_yolu)
        

    def islemler(self):
        if hasattr(self, 'ekran2') and self.ekran2:
            try:
                if self.ekran2.winfo_exists():
                    self.ekran2.destroy()
            except tk.TclError:
                pass  # Eğer widget zaten yoksa, hata alınmaz

        self.ekran4=Tk()
        self.ekran4.title("Klinik Adı")
        self.ekran4.configure(bg="#666666")
        self.ekran4.geometry("750x300+650+300")
        self.icon_yolu = self.get_resource_path("Klinik Logo resmi")
        
        try:
            # İkonu yükle
            self.ekran1.iconbitmap(self.icon_yolu)
        except Exception as e:
            print(f"Error loading icon: {e}")
        self.style = ttk.Style()
        self.style.configure("RoundedButton.TButton",
                        relief="flat",
                        background="#666666",
                        font=("Arial", 29),
                        borderwidth=1,
                        width=15)
        self.glr_arayuz_button = ttk.Button(self.ekran4, text="GELİR SAYFASI", style="RoundedButton.TButton",command=self.gelir_sayfası)
        self.glr_arayuz_button.grid(row=0, column=0, padx=20, pady=50)
        self.gdr_arayuz_button=ttk.Button(self.ekran4, text="GİDER SAYFASI", style="RoundedButton.TButton",command=self.gider_sayfasi)
        self.gdr_arayuz_button.grid(row=0, column=1, padx=20, pady=10)
        self.excel_arayuz_button=ttk.Button(self.ekran4, text="EXCEL DOSYASI", style="RoundedButton.TButton",command=self.bütün_excel_goster)
        self.excel_arayuz_button.grid(row=1, column=0, padx=20, pady=10)
        self.ciro_hesapla=ttk.Button(self.ekran4, text="CİRO HESAPLA", style="RoundedButton.TButton",command=self.aylikciro)
        self.ciro_hesapla.grid(row=1, column=1, padx=20, pady=10)
        self.ekran4.mainloop()

    def gelir_sayfası(self):
        if hasattr(self, 'ekran4') and self.ekran4:  
            try:  
                if self.ekran4.winfo_exists():  
                    self.ekran4.destroy()  # Ekran 4'ü kapat  
            except tk.TclError:  
                pass  # Eğer widget zaten yoksa, hata alınmaz  
        self.ekran5 = Tk()
        self.ekran5.title("Klinik Adı")
        self.ekran5.geometry("1500x750+350+100")
        self.ekran5.configure(bg="#666666")
        self.icon_yolu = self.get_resource_path("Klinik Logo resmi")
        
        try:
            # İkonu yükle
            self.ekran1.iconbitmap(self.icon_yolu)
        except Exception as e:
            print(f"Error loading icon: {e}")
        

        self.dosya_no=Label(text="DOSYA NO:", font=("Arial", 20), bg="#666666", fg="gold")
        self.dosya_no.place(x=100,y=75)
        self.dosya_no_grs=Entry(font=("Arial", 20))
        self.dosya_no_grs.place(x=300,y=75)

        self.gelir_ekle_lbl = Label(text="HASTA İSİM:", font=("Arial", 20), bg="#666666", fg="gold")
        self.gelir_ekle_lbl.place(x=100, y=150)
        self.gelir_ekle_entry = Entry(font=("Arial", 20))
        self.gelir_ekle_entry.place(x=300, y=150)

        self.label3 = Label(text="V1.7", fg="white", font='Arial', background="#666666")
        self.label3.pack(expand=1,anchor='se')

        self.gelir_ekle_tutar_lbl = Label(
            text="İŞLEM TUTARI:", font=("Arial", 20), bg="#666666", fg="gold")
        self.gelir_ekle_tutar_lbl.place(x=100, y=220)
        self.gelir_ekle_tutar_entry = Entry(font=("Arial", 20))
        self.gelir_ekle_tutar_entry.place(x=300, y=220)

        # Entry'ye her tuş basıldığında formatlama yap
        self.gelir_ekle_tutar_entry.bind("<KeyRelease>", self.format_value)
        """"
        self.checkbox_var=IntVar()
        self.checkbox = Checkbutton(self.ekran5,text="TAKSİT YAPMAK İSTİYORUM",font=("Arial", 10),bg="#666666",fg="gold",variable=self.checkbox_var,command=self.kontrol_et)

        self.checkbox.place(x=500, y=300)

        self.checkbox_pesin_var=IntVar()
        self.checkbox_pesin=Checkbutton(self.ekran5,text="PEŞİN ÖDEMEK İSTİYORUM",font=("Arial", 10),bg="#666666",fg="gold",variable=self.checkbox_pesin_var,command=self.kontrol_et2)
        self.checkbox_pesin.place(x=500,y=265)
        """
      
        self.ödenen=Label(text="ÖDENEN TUTARI:", font=("Arial", 20), bg="#666666", fg="gold")
        self.ödenen.place(x=60,y=320)
        self.ödenen_entry=Entry(font=("Arial", 20))
        self.ödenen_entry.place(x=300,y=320)
        self.ödenen_entry.bind("<KeyRelease>", self.format_number)
       



        self.doktor_uyar_label=Label(text="DOKTOR SEÇİNİZ", font=("Arial", 12), bg="#666666", fg="gold")
        self.doktor_uyar_label.place(x=1150,y=50)
        self.doktor_secenekler=["Doktor1","Doktor2","Doktor3","Doktor4","Doktor5"]
        self.doktor_secenekler_var=tk.StringVar(value=self.doktor_secenekler[0])
        self.doktor_secenekler_var.trace_add("write", self.doktor_degisti)
        
        self.doktor_secenekler_option=tk.OptionMenu(self.ekran5, self.doktor_secenekler_var, *self.doktor_secenekler)
        self.doktor_secenekler_option.place(x=1150,y=75)
        self.doktor_secenekler_option.configure(font=("Arial", 12), bg="#666666", fg="gold")

        self.gelir_goster_sayfa=Button(self.ekran5, text="GELİR SORGULA VE SİL",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.gelir_sorgu)
        self.gelir_goster_sayfa.place(x=275,y=450)
        
        self.gelir_temizle=Button(self.ekran5, text="TEMİZLE",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.gelir_ekle_temizle)
        self.gelir_temizle.place(x=630,y=450)
        """
        self.taksit_ekle=Button(self.ekran5, text="TAKSİT EKLE",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.taksit_fonk)
        self.taksit_ekle.place(x=650,y=450)
        """

        self.geridon_sayfa=Button(self.ekran5, text="GERİ DÖN",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.geridon_sayfaya)
        self.geridon_sayfa.place(x=800,y=450)
        """""
        self.listbox = Listbox(self.ekran5,width=40,height=6,font=("Arial", 20))
        self.listbox.place_forget()  
        self.listbox.bind("<<ListboxSelect>>", self.secim_yapildi) 
        """ 
        self.ekle_buton=Button(self.ekran5, text="GELİR EKLE",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.veri_grs)
        self.ekle_buton.place(x=50,y=450)

        self.alt_kategoriler = {
            "GÜLÜŞ TASARIM": ["ZİRKONYUM", "E-MAX", "LAMİNATE(YAPRAK PORSELEN)", "MONOLİTİK ZİRKONYUM","PORSELEN(METAL DESTEKLİ PROTEZ)","Tİ BASE MULTİ UNİT PORSELEN","Tİ BASE MULTİ ÜNİT MONOLİTİK ZİRKON","Tİ-BASE","Ti-BAR"],
            "PROTEZ": ["TEK ÇENE TOTAL PROTEZ", "TEK ÇENE HAREKETLİ İSKELET PROTEZ", "TEK ÇENE HİBRİT PROTEZ+BAR","TEK ÇENE HİBRİT PROTEZ+TİTANYUM BAR","OVERDENTURE PROTEZ","GECE PLAĞI","MULTİ UNİT"],
            "RESTO": ["ESTETİK DOLGU", "BONDİNG", "DOLGU","KUAFAJ","DİESTEMA KAPAMA","HASSASİYET GİDERİCİ","MASSETER BOTOKS+GECE PLAĞI","KANAL TEDAVİSİ+DOLGU","LAZER BLEACHİNG+DİŞ TAŞI TEMİZLİĞİ","KANAL YENİLEME"],
            "ÇOCUK": ["KANAL", "YER TUTUCU", "SÜT DİŞ ÇEKİMİ","DOLGU KOMPOMER","AMPUTASYON"],
            "CERRAHİ": ["DİŞ ÇEKİM","KOMPLİKE DİŞ ÇEKİM","SÜRMÜŞ YİRMİLİK DİŞ ÇEKİM","YARI GÖMÜLÜ DİŞ ÇEKİMİ","TAM GÖMÜLÜ DİŞ ÇEKİM","DİSTOANGULER GÖMÜLÜ DİŞ ÇEKİM","HORİZONTAL GÖMÜLÜ DİŞ ÇEKİM","KİSTEKTOMİ(KİST TEMİZLİĞİ)"],
            "PERİDONTOLOJİ": ["DETPOL", "TEK SEANS KÜRETAJ","ÇİFT SEANS KÜRETAJ","UZMAN KÜRETAJ","FRENEKTOMİ(UZMAN)","FLAP KÜRETAJ(UZMAN)","SDG(SERBEST DİŞETİ GREFT DİŞ BAŞI)"],
            "İMPLANT": ["BİLİM İMPLANT(YERLİ)","MEDİGMA(ALMAN)","OSSTEM(GÜNEY KORE)","MEDENTİKA(ALMAN)","STRAUMANN(İSVİÇRE)","AÇIK SİNÜS AMELİYATI(TEK TARAF)","KAPALI SİNÜS AMELİYATI","KEMİK TOZU(1cc)","MEMBRAN"]
        }

        # Seçim durumlarını saklamak için
        self.secenek_durumlari = {kategori: [tk.BooleanVar(value=False) for _ in alt] for kategori, alt in self.alt_kategoriler.items()}

        self.secenekler = list(self.alt_kategoriler.keys())

        # Dropdown menüsü
        self.val = tk.StringVar()
        self.val.set("YAPILACAK İŞLEM SEÇİNİZ")
        self.islem_secenek = tk.OptionMenu(self.ekran5, self.val, *self.secenekler, command=self.islem_secenekleri)
        self.islem_secenek.place(x=750, y=75)
        self.islem_secenek.configure(font=("Arial", 13), bg="#666666", fg="gold")

        self.checkbuttons = []  # Mevcut Checkbutton'ları saklamak için
    def format_value(self, event):
        raw_value = self.gelir_ekle_tutar_entry.get().replace(".", "")  
        if raw_value.isdigit():  
            value = int(raw_value) 
            self.formatted_value = "{:,}".format(value).replace(",", ".")  
            self.gelir_ekle_tutar_entry.delete(0, END)  # Eski değeri sil
            self.gelir_ekle_tutar_entry.insert(0, self.formatted_value) 
    def format_number(self,event):
        
        raw_value = self.ödenen_entry.get().replace(".", "")  
        if raw_value.isdigit():  
            value = int(raw_value)  #
            formatted_value = "{:,}".format(value).replace(",", ".")  

            # Entry'yi güncelle
            self.ödenen_entry.delete(0, END)  
            self.ödenen_entry.insert(0, formatted_value)  


    def islem_secenekleri(self, deger):
        # Eski Checkbutton'ları temizle
        for check in self.checkbuttons:
            check.destroy()
        self.checkbuttons.clear()

        # Seçilen kategoriye ait alt kategoriler
        alt_kategoriler = self.alt_kategoriler.get(deger, [])
        durumlar = self.secenek_durumlari.get(deger, [])

        # Yeni Checkbutton'ları oluştur
        for x, alt_kategori in enumerate(alt_kategoriler):
            var = durumlar[x]  # Var olan BooleanVar'ı kullan
            check = tk.Checkbutton(
                self.ekran5,
                text=alt_kategori,
                variable=var,
                font=("Arial", 13, "bold"),  # Yazı tipi
                bg="#666666",  # Arka plan rengi
                fg="gold",  # Yazı rengi
                selectcolor="black",  # Seçili alanın rengi
                activebackground="#cccccc",  # Hover arka plan rengi
                activeforeground="blue"  # Hover yazı rengi
            )
            check.place(x=750, y=120 + x * 30)
            self.checkbuttons.append(check)
        
    def secilen_degerleri_al(self):
        self.secilen_degerler = []

        # Seçenek durumlarını kontrol et
        for kategori, durumlar in self.secenek_durumlari.items():
            alt_kategoriler = self.alt_kategoriler[kategori]

            # Durumları kontrol et ve seçilenleri ekle
            for i, durum in enumerate(durumlar):
                if durum.get():  # Eğer True ise
                     self.secilen_degerler.append(alt_kategoriler[i])

        return  self.secilen_degerler
    """
    def seçenekler(self):  
        self.uyarı = None  
        self.gelir = self.gelir_ekle_tutar_entry.get()  
        self.durum = True  
        
        if not self.gelir:  
            self.uyarı = Label(text="0 TL TAKSİTLENDİRİLEMEZ", font=("Arial", 25), bg="#666666", fg="gold")  
            self.uyarı.place(x=100, y=350)  
            self.durum = False  
        else:  
            # Binlik ayırıcıları kaldırma  
            self.gelir = self.gelir.replace('.', '').replace(',', '.')  # '.' karakterini kaldır, gerekirse ',' ile '.' değiştirilir  
            self.gelir = int(self.gelir)  # Sonucu float'a dönüştür  

            # Seçenekler  
            self.option_list = [  
                (f"2 AY TAKSİT ({(self.gelir/2)} TL)", self.gelir),  
                (f"3 AY TAKSİT ({(self.gelir/3)} TL)", self.gelir),  
                (f"4 AY TAKSİT ({self.gelir/4} TL)", self.gelir),  
                (f"6 AY TAKSİT ({self.gelir/6} TL)", self.gelir),  
                (f"8 AY TAKSİT ({self.gelir/8} TL)", self.gelir),  
                (f"12 AY TAKSİT ({self.gelir/12} TL)", self.gelir)  
            ]  
            
            for option in self.option_list:  
                self.listbox.insert(END, option[0])   
        
        return self.durum

    def secim_yapildi(self, event=None):  
        self.taksit_adet = 0  # Başlangıçta taksit adedini sıfırla  
        self.secilen = self.listbox.curselection()  
        
        if self.secilen:  # Eğer bir seçim varsa  
            self.index_no = self.secilen[0]  # Seçilen indeks  

            # Seçilen taksit seçeneğinin metnini al  
            self.selected_option = self.option_list[self.index_no][0]  

            # Seçilen taksit seçeneğine göre taksit adedini belirleme  
            if "2 AY TAKSİT" in self.selected_option:  
                self.taksit_adet = 2  
                self.tk_tutari=self.gelir/2
            elif "3 AY TAKSİT" in self.selected_option:  
                self.taksit_adet = 3  
                self.tk_tutari=self.gelir/3
            elif "4 AY TAKSİT" in self.selected_option:  
                self.taksit_adet = 4  
                self.tk_tutari=self.gelir/4
            elif "6 AY TAKSİT" in self.selected_option:  
                self.taksit_adet = 6  
                self.tk_tutari=self.gelir/6
            elif "8 AY TAKSİT" in self.selected_option:  
                self.taksit_adet = 8  
                self.tk_tutari=self.gelir/8
            if "12 AY TAKSİT" in self.selected_option:  # Burayı düzelttik  
                self.taksit_adet = 12  
                self.tk_tutari=self.gelir/12


    def kontrol_et(self):
        if self.checkbox_var.get() == 1:
            self.durumu=self.seçenekler() 
            self.checkbox_pesin_var.set(0)
            if self.durumu==True:            
                self.listbox.place(x=100, y=350)  
                self.ekran5.configure(bg="#666666")
                self.ekle_buton.place(x=50,y=575)
                self.gelir_goster_sayfa.place(x=275,y=575)
                self.taksit_ekle.place(x=650,y=575)
                self.gelir_temizle.place(x=875,y=575)
                self.geridon_sayfa.place(x=1050,y=575)
        else:
            self.listbox.delete(0, END)
            self.listbox.place_forget() 
            self.ekran5.configure(bg="#666666")
            self.ekle_buton.place(x=50,y=450)
            self.gelir_goster_sayfa.place(x=275,y=450)
            self.gelir_temizle.place(x=875,y=450)
            self.taksit_ekle.place(x=650,y=450)
            self.geridon_sayfa.place(x=1050,y=450)
            if hasattr(self, "durum"):
                if self.durum == False:
                    self.uyarı.destroy()

    def kontrol_et2(self):
        if self.checkbox_pesin_var.get() == 1:
            self.checkbox_var.set(0)
        if self.checkbox_var.get() == 0:
            self.listbox.delete(0, END)
            self.listbox.place_forget()
            self.ekran5.configure(bg="#666666")
            self.ekle_buton.place(x=50, y=450)
            self.gelir_goster_sayfa.place(x=275, y=450)
            self.gelir_temizle.place(x=875, y=450)
            self.taksit_ekle.place(x=650,y=450)
            self.geridon_sayfa.place(x=1050, y=450)
            
            
            if hasattr(self, 'uyarı') and self.uyarı is not None:
                self.uyarı.destroy()
                
    """
    def gelir_ekle_temizle(self):
        """
        self.checkbox_pesin_var.set(0)
        self.checkbox_var.set(0)
        self.kontrol_et()
        """
        for durum_listesi in self.secenek_durumlari.values():
            for var in durum_listesi:
                var.set(False)

        # Ekrandaki tüm Checkbutton'ları temizle
        for check in self.checkbuttons:
            check.destroy()
        self.checkbuttons.clear()
        self.entries=[self.dosya_no_grs,self.gelir_ekle_entry,self.gelir_ekle_tutar_entry,self.ödenen_entry]
        for entry in self.entries:
            entry.delete(0,tk.END)

    def geridon_sayfaya(self):
        self.ekran5.destroy()
        self.islemler()
    def doktor_degisti(self, *args):
        self.doktor = self.doktor_secenekler_var.get()
        if self.doktor=="Doktor4":
                self.prim_lblms=Label(text="DOKTOR PRİMİ GİRİNİZ:", font=("Arial", 12), bg="#666666", fg="gold")
                self.prim_lblms.place(x=1150,y=150)
                self.prim_entryms=Entry(font=("Arial", 20))
                self.prim_entryms.place(x=1150,y=175)
        elif self.doktor=="Doktor5":
                self.prim_lblbe=Label(text="DOKTOR PRİMİ GİRİNİZ:", font=("Arial", 12), bg="#666666", fg="gold")
                self.prim_lblbe.place(x=1150,y=150)
                self.prim_entrybe=Entry(font=("Arial", 20))
                self.prim_entrybe.place(x=1150,y=175)
        else:
            if hasattr(self, "prim_entrybe") and self.prim_entrybe.winfo_ismapped():
                self.prim_entrybe.place_forget()
                self.prim_lblbe.place_forget()
            elif hasattr(self, "prim_entryms") and self.prim_entryms.winfo_ismapped():
                self.prim_entryms.place_forget()
                self.prim_lblms.place_forget()


    def veri_grs(self):
        self.satir_sayisi=2267
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)
        self.sayfa=self.excel_dosya["GELİR"]
        self.brüt_sayfa=self.excel_dosya["NET GELİR"]
        self.sayfa_gider=self.excel_dosya["GİDER"]
        self.bb=True
        self.gunun_tarihi=datetime.now().strftime("%d.%m.%Y")
        self.dosya_no_grs_degeri=self.dosya_no_grs.get()
        self.gelir_ekle_entry_degeri=self.gelir_ekle_entry.get().upper()
        self.gelir_ekle_tutar_entry_degeri=self.gelir_ekle_tutar_entry.get()
        self.ödenen_tutar=self.ödenen_entry.get()
        self.secilen_degerler=self.secilen_degerleri_al()
        if not self.dosya_no_grs_degeri:
            tk.messagebox.showerror("Hata","DOSYA NO BOŞ OLAMAZ")
            self.bb=False
        elif not self.gelir_ekle_entry_degeri:
            tk.messagebox.showerror("Hata","GELİR İSMİ BOŞ OLAMAZ")
            self.bb=False
        elif not self.gelir_ekle_tutar_entry_degeri:
            tk.messagebox.showerror("Hata","GELİR TUTARI BOŞ OLAMAZ")
            self.bb=False
        elif    not self.ödenen_tutar:
            tk.messagebox.showerror("Hata","ÖDENEN TUTAR BOŞ OLAMAZ")
            self.bb=False
        """
        elif self.checkbox_var.get()==0 and self.checkbox_pesin_var.get()==0:
           tk.messagebox.showerror("Hata","PEŞİN VEYA TAKSİT SEÇENEĞİNİ SEÇMEDİNİZ")
            self.bb=False
        """

        raw_value = self.gelir_ekle_tutar_entry.get().replace(".", "")  # Noktaları kaldır
        if raw_value.isdigit():  # Değerin sayı olup olmadığını kontrol et
            self.gelir_ekle_tutar_entry_degeri = int(raw_value)  # Değeri integer olarak al
        if self.bb==True:
            while True:
                # 15 sütuna kadar kontrol et
                self.veri_varmi = any(self.sayfa.cell(row=self.satir_sayisi, column=col).value is not None for col in range(1, 16))

                if not self.veri_varmi:  
                    break
                self.satir_sayisi += 1 
            self.sayfa.cell(row=self.satir_sayisi, column=1).value = self.gunun_tarihi  # Değeri ata
            self.sayfa.cell(row=self.satir_sayisi, column=2).value = self.dosya_no_grs_degeri
            self.sayfa.cell(row=self.satir_sayisi, column=3).value = self.gelir_ekle_entry_degeri
            self.sayfa.cell(row=self.satir_sayisi, column=4).value = ", ".join(self.secilen_degerler)
            self.gider=self.islem_gider()
            self.doktor_prim=self.doktor_secim()
            self.doktorsecim=self.doktor_secenekler_var.get()
            self.brütgelir=0
            self.max_satir = self.sayfa_gider.max_row
            self.gider_satir_sayisi=1
            while True:
                veri_varmi = self.sayfa_gider.cell(row=self.gider_satir_sayisi, column=1).value
                if veri_varmi is None:
                    break
                self.gider_satir_sayisi += 1
 
            if self.doktorsecim=="Doktor1":
                self.sayfa.cell(row=self.satir_sayisi, column=7).value =self.doktor_prim 
                self.brütgelir=int(self.gelir_ekle_tutar_entry_degeri)-int(self.doktor_prim+int(self.gider))
                self.sayfa.cell(row=self.satir_sayisi, column=15).value = self.brütgelir
                self.sayfa_gider.cell(row=self.gider_satir_sayisi,column=4).value=self.doktor_prim
            elif self.doktorsecim=="Doktor2":
                self.sayfa.cell(row=self.satir_sayisi, column=8).value =self.doktor_prim 
                self.brütgelir=int(self.gelir_ekle_tutar_entry_degeri)-int(self.doktor_prim+int(self.gider))
                self.sayfa.cell(row=self.satir_sayisi, column=15).value = self.brütgelir
                self.sayfa_gider.cell(row=self.gider_satir_sayisi,column=5).value=self.doktor_prim
            elif self.doktorsecim=="Doktor3":
                self.sayfa.cell(row=self.satir_sayisi, column=9).value =self.doktor_prim 
                self.brütgelir=int(self.gelir_ekle_tutar_entry_degeri)-int(self.doktor_prim+int(self.gider))
                self.sayfa.cell(row=self.satir_sayisi, column=15).value = self.brütgelir
                self.sayfa_gider.cell(row=self.gider_satir_sayisi,column=6).value=self.doktor_prim
            elif self.doktorsecim=="Doktor4":
                self.prim_entryms_get=self.prim_entryms.get()
                self.sayfa.cell(row=self.satir_sayisi, column=10).value =self.prim_entryms_get 
                self.brütgelir=int(self.gelir_ekle_tutar_entry_degeri)-(int(self.prim_entryms_get)+self.gider)
                self.sayfa.cell(row=self.satir_sayisi, column=15).value = self.brütgelir
                self.sayfa_gider.cell(row=self.gider_satir_sayisi,column=7).value=self.prim_entryms_get
            elif self.doktorsecim=="Doktor5":
                self.prim_entrybe_get=self.prim_entrybe.get()
                self.sayfa.cell(row=self.satir_sayisi, column=11).value =self.prim_entrybe_get 
                self.brütgelir=int(self.gelir_ekle_tutar_entry_degeri)-(int(self.prim_entrybe_get)+self.gider)
                self.sayfa.cell(row=self.satir_sayisi, column=15).value = self.brütgelir
                self.sayfa_gider.cell(row=self.gider_satir_sayisi,column=8).value=self.prim_entrybe_get
            
            self.sayfa.cell(row=self.satir_sayisi, column=5).value = self.gider

            self.sayfa.cell(row=self.satir_sayisi, column=13).value = self.ödenen_tutar

            self.sayfa.cell(row=self.satir_sayisi, column=6).value = "NAKİT"

            """
            if self.checkbox_pesin_var.get()==1:
                self.sayfa.cell(row=self.satir_sayisi, column=6).value = "PEŞİN"
                self.sayfa.cell(row=self.satir_sayisi, column=13).value = f"₺ {self.gelir_ekle_tutar_entry_degeri}"
                self.sayfa.cell(row=self.satir_sayisi, column=14).value ="₺ 0.00"
                
            else:
                self.br=int(self.tk_tutari) 
                self.kalan=int(self.gelir_ekle_tutar_entry_degeri-self.tk_tutari)
                self.sayfa.cell(row=self.satir_sayisi, column=6).value =self.selected_option
                self.sayfa.cell(row=self.satir_sayisi, column=14).value =f"₺ {(self.kalan)}"
                self.sayfa.cell(row=self.satir_sayisi, column=13).value = f"₺{(self.br)}"
                self.sayfa.cell(row=self.satir_sayisi, column=15).value =self.taksit_adet
                self.sayfa.cell(row=self.satir_sayisi, column=16).value =(self.taksit_adet-1)
            """
            self.sayfa.cell(row=self.satir_sayisi, column=12).value = f"₺ {self.gelir_ekle_tutar_entry_degeri}"
            self.kalan = int(self.gelir_ekle_tutar_entry_degeri) - int(self.ödenen_tutar.replace(".", ""))
            self.sayfa.cell(row=self.satir_sayisi, column=14).value = f"₺ {self.kalan}"


            self.sayfa_gider.cell(row=self.gider_satir_sayisi,column=1).value=self.gunun_tarihi
            self.sayfa_gider.cell(row=self.gider_satir_sayisi,column=2).value=", ".join(self.secilen_degerler)
            self.sayfa_gider.cell(row=self.gider_satir_sayisi,column=3).value=self.gider
            self.excel_dosya.save( self.excel_dosya_yolu) 
            try:
                self.excel_dosya.save(self.excel_dosya_yolu)
            except Exception as e:
                messagebox.showerror("Hata", f"{e}")

            oku= self.sayfa.cell(row=self.satir_sayisi,column=2).value
            messagebox.showinfo("Bilgi", "İşlem başarıyla tamamlandı!")
            self.gelir_ekle_temizle()
            

    def islem_gider(self):
        self.islem_fiyatları_txt=self.get_resource_path(r"klinik islem ücretleri.txt")
        self.islem_gider_toplam=0
        for islem in self.secilen_degerler:
            with open(self.islem_fiyatları_txt, "r", encoding="utf-8") as dosya:
                for satir in dosya:
                    sutunlar = satir.strip().split("=")
                    if islem in sutunlar:
                        deger=sutunlar[1]
                        deger=int(deger)
                        self.islem_gider_toplam+=deger
        return self.islem_gider_toplam
            

        




    
    def doktor_secim(self):
            self.maas=0
            self.doktor=self.doktor_secenekler_var.get()
            self.brüt=int(self.gelir_ekle_tutar_entry_degeri-self.gider)
            if self.doktor=="Doktor1":
                self.maas=int((self.brüt*20)/100)
            elif self.doktor=="Doktor2":
                self.maas=int((self.brüt*15)/100)
            elif self.doktor=="Doktor3":
                self.maas=int((self.brüt*30)/100)
            return self.maas

    """
    def taksit_fonk(self):
        
        self.taksit_ekran = Tk()
        self.satir_sayisi=2268 
        self.taksit_ekran.title("Klinik Adı")
        self.taksit_ekran.geometry("750x750+550+150")
        self.taksit_ekran.configure(bg="#666666")
        self.taksit_ekran.iconbitmap("images.ico")
        self.taksit_ekran.attributes("-topmost", True)

        self.taksit_gelir_lbl=Label(self.taksit_ekran,text="HASTA İSMİ VE SOYİSMİ GİRİNİZ", font=("Arial", 20), bg="#666666", fg="gold")
        self.taksit_gelir_lbl.place(x=25,y=100)

        self.taksit_gelir_tutar=Entry(self.taksit_ekran,font=("Arial", 20))
        self.taksit_gelir_tutar.place(x=255,y=200)

        self.kk=Label(self.taksit_ekran,text="", font=("Arial", 20), bg="#666666", fg="gold")
        self.kk.place(x=100,y=450)
        
        
        self.taksit_buton=Button(self.taksit_ekran, text="TAKSİT EKLE",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.taksit)
        self.taksit_buton.place(x=250,y=300)
    def taksit(self):
        self.taksit_tutari_get = self.taksit_gelir_tutar.get().strip().upper()
        self.gunun_tarihi = datetime.now().strftime("%d.%m.%Y")
        self.max_satir = self.sayfa.max_row
        kalan_taksit_endüşük = float("inf")  # Başlangıç değeri
        satir_en_dusuk = None  # En düşük kalan taksiti tutacak değişken

        if not self.taksit_tutari_get:
            messagebox.showerror("Hata", "Lütfen hasta ismini eksiksiz giriniz!")
            return

        for satir in range(self.satir_sayisi, self.max_satir + 1):
            isim = self.sayfa.cell(satir, 3).value
            isim = str(isim).strip()
            kalan_taksit = self.sayfa.cell(satir, 16).value

            if isinstance(kalan_taksit, (int, float)) and self.taksit_tutari_get == isim and kalan_taksit > 0:
                # En düşük kalan taksiti bul
                if kalan_taksit < kalan_taksit_endüşük:
                    kalan_taksit_endüşük = kalan_taksit
                    satir_en_dusuk = satir  # Bu satırı en düşük kalan taksitle sakla

        # Eğer en düşük kalan taksitli satır bulunamazsa hata ver
        if satir_en_dusuk is None:
            messagebox.showerror("Hata", "İlgili hasta bulunamadı veya taksit bilgisi eksik!")
            return

        # En düşük kalan taksite sahip satır bulundu
        satir = satir_en_dusuk
        isim = self.sayfa.cell(satir, 3).value
        kalan_taksit = self.sayfa.cell(satir, 16).value
        dosya_no = self.sayfa.cell(satir, 2).value
        yapılan_islem = self.sayfa.cell(satir, 4).value
        taksit_tutari = self.sayfa.cell(satir, 6).value
        sütün = 7
        doktor = None

        for i in range(7, 12):
            if self.sayfa.cell(satir, i).value is not None:
                doktor = self.sayfa.cell(satir, i).value
                sütün = i  

        toplam_ücret = self.sayfa.cell(satir, 12).value
        ödenen = self.sayfa.cell(satir, 13).value
        kalan = self.sayfa.cell(satir, 14).value
        taksit_adeti = self.sayfa.cell(satir, 15).value
        kalan_taksit_endüşük -= 1  # Kalan taksiti azalt

        while True:
            self.veri_varmi = any(
                self.sayfa.cell(row=self.satir_sayisi, column=col).value is not None
                for col in range(1, 16)
            )
            if not self.veri_varmi:
                break
            self.satir_sayisi += 1

        self.sayfa.cell(row=self.satir_sayisi, column=1).value = self.gunun_tarihi
        self.sayfa.cell(row=self.satir_sayisi, column=2).value = dosya_no
        self.sayfa.cell(row=self.satir_sayisi, column=3).value = self.taksit_tutari_get
        self.sayfa.cell(row=self.satir_sayisi, column=4).value = yapılan_islem
        self.sayfa.cell(row=self.satir_sayisi, column=6).value = taksit_tutari
        self.sayfa.cell(row=self.satir_sayisi, column=sütün).value = doktor
        self.sayfa.cell(row=self.satir_sayisi, column=12).value = toplam_ücret
        self.sayfa.cell(row=self.satir_sayisi, column=13).value = ödenen
        self.sayfa.cell(row=self.satir_sayisi, column=14).value = kalan
        self.sayfa.cell(row=self.satir_sayisi, column=15).value = taksit_adeti
        self.sayfa.cell(row=self.satir_sayisi, column=16).value = kalan_taksit_endüşük

        self.kk.configure(text=f"DOSYA NO: {dosya_no}\n | İSİM: {isim} |\n KALAN TAKSİT ADETİ: {kalan_taksit_endüşük}|\n TAKSİT ADETİ: {taksit_adeti}|\n TAKSİT TUTARI: {taksit_tutari}")
        messagebox.showinfo("Bilgi", "İşlem başarıyla tamamlandı!")
        self.excel_dosya.save("excel tablo yolu ")

                    
                    
         
    """  
    
    def gelir_sorgu(self):
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)
        self.sayfa=self.excel_dosya["GELİR"]
        self.brüt_sayfa=self.excel_dosya["NET GELİR"]
        self.sayfa_gider=self.excel_dosya["GİDER"]
        self.ekran_gelir_sorgu_ekran = Tk()
        self.ekran_gelir_sorgu_ekran.title("Klinik Adı")
        self.ekran_gelir_sorgu_ekran.geometry("1500x950+300+20")
        self.ekran_gelir_sorgu_ekran.configure(bg="#666666")
        self.icon_yolu = self.get_resource_path("Klinik Logo resmi")
        
        try:
            # İkonu yükle
            self.ekran1.iconbitmap(self.icon_yolu)
        except Exception as e:
            print(f"Error loading icon: {e}")
        self.ekran_gelir_lbl=Label(self.ekran_gelir_sorgu_ekran,text="HASTA İSMİ SOYİSMİ GİRİNİZ:", font=("Arial", 20), bg="#666666", fg="gold")
        self.ekran_gelir_lbl.place(x=250,y=100)
        self.ekran_gelir_entry=Entry(self.ekran_gelir_sorgu_ekran,font=("Arial", 20))
        self.ekran_gelir_entry.place(x=650,y=100)
        self.ekran_gelir_btn=Button(self.ekran_gelir_sorgu_ekran, text="SORGULA",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.gelir_sorgu_fonk)
        self.ekran_gelir_btn.place(x=350,y=300)

        self.ödeme_ekle_lbl=Label(self.ekran_gelir_sorgu_ekran,text="ÖDEME GİRİNİZ: ", font=("Arial", 20), bg="#666666", fg="gold")
        self.ödeme_ekle_lbl.place(x=250,y=200)
        self.ödeme_ekle_entry=Entry(self.ekran_gelir_sorgu_ekran,font=("Arial", 20))
        self.ödeme_ekle_entry.place(x=500,y=200)

        self.ekran_gelir_sorgu_ekle_btn=Button(self.ekran_gelir_sorgu_ekran, text="ÖDEME EKLE",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.ödeme_ekle)
        self.ekran_gelir_sorgu_ekle_btn.place(x=550,y=300)
        
        self.ekran_gelir_sil_btn=Button(self.ekran_gelir_sorgu_ekran, text="GELİR SİL",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.gelir_sil)
        self.ekran_gelir_sil_btn.place(x=800,y=300)

        self.ekran_gelir_tree=ttk.Treeview(self.ekran_gelir_sorgu_ekran, columns=("Tarih", "Dosya No", "İsim", "Yapılan İşlem", "İşlem Tutarı", "Ödenen", "Kalan"), show="headings")

        for col in self.ekran_gelir_tree["columns"]:
            self.ekran_gelir_tree.heading(col, text=col)  # Başlıkları ayarlama

            # Koşula göre genişlik ayarlama
            if col == "Tarih" or col == "Dosya No":  
                self.ekran_gelir_tree.column(col, width=100)  
            elif col == "Yapılan İşlem":  
                self.ekran_gelir_tree.column(col, width=350)  
            else:
                self.ekran_gelir_tree.column(col, width=200) 
        
        self.ekran_gelir_tree.place(x=50,y=400)
        self.ekran_gelir_tree.place_forget()

        self.ekran_gelir_sonuc_göter=Label(self.ekran_gelir_sorgu_ekran,text="", font=("Arial", 20), bg="#666666", fg="gold")
        self.ekran_gelir_sonuc_göter.place(x=250,y=550)
        self.ekran_gelir_sonuc_göter.place_forget()
        
    
    def gelir_sorgu_fonk(self):
        self.ekran_gelir_tree.delete(*self.ekran_gelir_tree.get_children())
        self.satir_sayisi=2267
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)
        self.sayfa=self.excel_dosya["GELİR"]
        self.brüt_sayfa=self.excel_dosya["NET GELİR"]
        self.sayfa_gider=self.excel_dosya["GİDER"]
        
        self.sorgulanan_kişiler=[]
        self.ekran_gelir_entry_get=self.ekran_gelir_entry.get().upper()
        max_satır=self.sayfa.max_row
        if  self.ekran_gelir_entry_get:
            self.ekran_gelir_sonuc_göter.place(x=350,y=750)
            
            for satir in range(self.satir_sayisi,max_satır+1):
                self.sorgu_kişi=[]
                sorgu_ismi=self.sayfa.cell(row=satir,column=3).value
                if sorgu_ismi==self.ekran_gelir_entry_get:
                    for i in range(1,4):
                        deger=self.sayfa.cell(row=satir,column=i).value
                        self.sorgu_kişi.append(deger)
                    deger=self.sayfa.cell(row=satir,column=4).value
                    self.sorgu_kişi.append([deger])
                    for j in range(12,15):
                        deger=self.sayfa.cell(row=satir,column=j).value
                        self.sorgu_kişi.append(deger)
                if self.sorgu_kişi:
                    self.sorgulanan_kişiler.append(self.sorgu_kişi)
                  
            if self.sorgulanan_kişiler:
                self.ekran_gelir_tree.place(x=50,y=400)
                for kişi in self.sorgulanan_kişiler:
                    self.ekran_gelir_tree.insert("",tk.END,values=kişi)
                for col in self.ekran_gelir_tree["columns"]:
                    self.ekran_gelir_tree.column(col, anchor='center')
                
            
                    #En küçük değerleri bulma7 değer var 6 indeksli
                    self.toplam_yapılan_işlem_ücreti=0
                    self.toplam_ödenen_para=0
                    self.kalan_para=0
                    i = 0
                    
                    sayı=len(self.sorgulanan_kişiler)

                    for kişi in self.sorgulanan_kişiler:
                        deger=kişi[5]
                        sayisal_deger = int(deger.replace('.', ''))  # Binlik ayracını kaldır ve integer'a çevir
                        self.toplam_ödenen_para+=sayisal_deger
                        toplam_ödenen_para = "{:,}".format(self.toplam_ödenen_para).replace(",", ".")

                    while i < len(self.sorgulanan_kişiler)-1 :
                        j=0
                        # Eğer ürünler aynıysa (4. sütun)
                        j=i+1
                        while j<sayı:
                            
                            if self.sorgulanan_kişiler[i][3] == self.sorgulanan_kişiler[j][3]:
                                fiyat1 = int(self.sorgulanan_kişiler[i][6].replace('₺', '').replace(' ', ''))
                                fiyat2 = int(self.sorgulanan_kişiler[j][6].replace('₺', '').replace(' ', ''))
                                if self.sorgulanan_kişiler[i][1] == self.sorgulanan_kişiler[j][1]:
                                    if fiyat1 > fiyat2:
                                        self.sorgulanan_kişiler.remove(self.sorgulanan_kişiler[i])
                                        
                                        sayı=len(self.sorgulanan_kişiler)
                                        j=i
                                    else:
                                        self.sorgulanan_kişiler.remove(self.sorgulanan_kişiler[j])
                                     
                                        sayı=len(self.sorgulanan_kişiler)
                                        j=i                          
                            j+=1
                    
                        i += 1
                    for kişi in self.sorgulanan_kişiler:
                        işem_ücret = int(kişi[4].replace('₺', '').replace(' ', ''))
                        kalan_para=int(kişi[6].replace('₺', '').replace(' ', ''))
                        self.toplam_yapılan_işlem_ücreti+=işem_ücret
                        
                    
                        self.kalan_para+=kalan_para
                    
                    toplam_işem = "{:,}".format(self.toplam_yapılan_işlem_ücreti).replace(",", ".")
                    kalan_tutar="{:,}".format(self.kalan_para).replace(",", ".")
                    gosterilecek_deger=f"TOPLAM İŞLEM TUTARI: {toplam_işem}\n TOPLAM ÖDENEN PARA: {toplam_ödenen_para}\n KALAN TUTAR: {kalan_tutar}"
                    self.ekran_gelir_sonuc_göter.configure(text=gosterilecek_deger)
            else:
                self.ekran_gelir_sonuc_göter.configure(text="KİŞİ BULUNAMADI")
        else:
            messagebox.showerror("HATA","SORGULANACAK KİŞİ GİRİNİZ")
        



    def ödeme_ekle(self):
        self.satir_sayisi=2267
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)
        self.sayfa=self.excel_dosya["GELİR"]
        self.brüt_sayfa=self.excel_dosya["NET GELİR"]
        self.sayfa_gider=self.excel_dosya["GİDER"]
        self.ödeme_ekle_entry_get=self.ödeme_ekle_entry.get()
        self.ekran_gelir_entry_get=self.ekran_gelir_entry.get().upper()
        self.kişiler=[]
        self.doktor_satır=[]
        max_satır=self.sayfa.max_row
        if self.ödeme_ekle_entry_get:
            for satir in range(self.satir_sayisi,max_satır+1):
                self.kişi=[]
                deger=self.sayfa.cell(row=satir,column=3).value
                if deger:
                    if self.ekran_gelir_entry_get==deger:
                        for i in range(1,7):
                            if i==4:
                                deger=self.sayfa.cell(row=satir,column=i).value
                                self.kişi.append([deger])
                            else:
                                deger=self.sayfa.cell(row=satir,column=i).value
                                self.kişi.append(deger)
                        for j in range(7,12):
                            deger=self.sayfa.cell(row=satir,column=j).value
                            if deger:
                                self.kişi.append(deger)
                                self.doktor_satır.append(j)
                        for k in range(12,15):
                            deger=self.sayfa.cell(row=satir,column=k).value
                            self.kişi.append(deger)
                
                        self.kişiler.append(self.kişi)
            
            #10 elemanlı 9 indexli liste
            i=0
            sayı=len(self.kişiler)
            while  i < len(self.kişiler) - 1:
                j=0
                # Eğer ürünler aynıysa (4. sütun)
                j=i+1
                while j<sayı:
                    if self.kişiler[i][3] == self.kişiler[j][3] and self.kişiler[i][0]==self.kişiler[j][0]:
                        fiyat1 = int(self.kişiler[i][9].replace('₺', '').replace(' ', ''))
                        fiyat2 = int(self.kişiler[j][9].replace('₺', '').replace(' ', ''))
                        
                    
                        if fiyat1 > fiyat2:
                            self.kişiler.pop(i) 
                            self.doktor_satır.pop(i)
                            sayı=len(self.kişiler)
                            j=i
                        else:
                            self.kişiler.pop(j)  
                            self.doktor_satır.pop(j)
                            sayı=len(self.kişiler)
                            j=i
                    
                    j+=1
                i += 1 
                        
            for i in range(len(self.kişiler) - 1, -1, -1):  # Ters yönde döngü  
                if self.kişiler[i][9] == '₺ 0':  
                    self.kişiler.pop(i)  # İndeksle öğeyi kaldır
        
            odenen_tutar=int(self.ödeme_ekle_entry_get)

            for i in range(len(self.kişiler)):
                deger = self.kişiler[i][9]  
                kalan = int(deger.replace('₺', '').replace(' ', '')) 

                if kalan > odenen_tutar:
                    sonuc = kalan - odenen_tutar
                    deger = f"₺ {sonuc}"
                    formatted_deger = "{:,}".format(odenen_tutar).replace(",", ".")  
                    self.kişiler[i][8] = formatted_deger
                    self.kişiler[i][9]=f"₺ {sonuc}"
                    if i + 1 < len(self.kişiler): 
                        self.kişiler.pop(i + 1)
                        break  

                elif kalan <= odenen_tutar:
                    odenen_tutar -= kalan
                    deger = "₺ 0"
                    formatted_deger = "{:,}".format(kalan).replace(",", ".")  
                    self.kişiler[i][8] = formatted_deger 
                    self.kişiler[i][9]=deger



            self.satir_sayisi=2268
            while True:
                if self.sayfa.cell(row=self.satir_sayisi, column=1).value not in [None, ""]:
                    self.satir_sayisi += 1
                else:
                    break
            index=0
            
            for kişi in self.kişiler:
                for i in range(1,7):
                    if i==4:
                        deger=kişi[3]
                        string=""
                        for deger1 in deger:
                            string = ",".join(str(deger1) for deger1 in deger)
                        self.sayfa.cell(row=self.satir_sayisi,column=i).value=string
                    else:
                        deger=kişi[i-1]
                        self.sayfa.cell(row=self.satir_sayisi,column=i).value=deger
                doktor=kişi[6]
                dongü=7
                self.sayfa.cell(row=self.satir_sayisi,column=self.doktor_satır[index]).value=doktor
                for j in range(12,15):
                    self.sayfa.cell(row=self.satir_sayisi,column=j).value=kişi[dongü]
                    dongü+=1
                self.satir_sayisi+=1
                index+=1
            self.excel_dosya.save(self.excel_dosya_yolu)
            self.ekran_gelir_sonuc_göter.configure(text="ÖDEME BAŞARIYLA EKLENDİ")
        else:
            messagebox.showerror("HATA","ÖDENECEK TUTAR GİRİNİZ")


    def gelir_sil(self):
            self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
            self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)
            self.sayfa=self.excel_dosya["GELİR"]
            self.brüt_sayfa=self.excel_dosya["NET GELİR"]
            self.sayfa_gider=self.excel_dosya["GİDER"]
            secilen = self.ekran_gelir_tree.focus()
            if secilen:
                degerler=self.ekran_gelir_tree.item(secilen,"values")
                degerler_dizi=list(degerler)
                self.satir_sayisi=2267
                max=self.sayfa.max_row
                for satır in range(self.satir_sayisi,max+1):
                    tarih=self.sayfa.cell(row=satır,column=1).value 
                    dosya_no=self.sayfa.cell(row=satır,column=2).value 
                    isim=self.sayfa.cell(row=satır,column=3).value 
                    islem=self.sayfa.cell(row=satır,column=4).value 
                    islem_tutar=self.sayfa.cell(row=satır,column=12).value 
                    ödenen=self.sayfa.cell(row=satır,column=13).value 
                    kalan=self.sayfa.cell(row=satır,column=14).value
                    if degerler_dizi[0]==tarih and  degerler_dizi[1]==dosya_no and degerler_dizi[2]==isim and degerler_dizi[3]==islem and degerler_dizi[4]==islem_tutar and degerler_dizi[5]==ödenen and degerler_dizi[6]==kalan:
                        self.sayfa.delete_rows(satır)
                self.excel_dosya.save(self.excel_dosya_yolu)
                messagebox.showinfo("BAŞARILI","BAŞARIYLA SİLİNDİ")
            else:
                messagebox.showerror("HATA","HASTA SEÇMEDİNİZ")
                


    def gider_sayfasi(self):
        if hasattr(self, 'ekran4') and self.ekran4:
            try:
                if self.ekran4.winfo_exists():
                    self.ekran4.destroy()
            except tk.TclError:
                pass 
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)
        self.sayfa=self.excel_dosya["GELİR"]
        self.brüt_sayfa=self.excel_dosya["NET GELİR"]
        self.sayfa_gider=self.excel_dosya["GİDER"]        
        self.ekran_gider=Tk()
        self.ekran_gider.title("Klinik Adı")
        self.ekran_gider.configure(bg="#666666")
        self.ekran_gider.geometry("1650x850+150+70")
        self.icon_yolu = self.get_resource_path("Klinik Logo resmi")
        
        try:
            # İkonu yükle
            self.ekran1.iconbitmap(self.icon_yolu)
        except Exception as e:
            print(f"Error loading icon: {e}")
        self.gider_ismi_lbl=Label(self.ekran_gider,text="GİDER İSMİ GİRİNİZ: ", font=("Arial", 20), bg="#666666", fg="gold")
        self.gider_ismi_lbl.place(x=5,y=100)
        self.gider__ismi_entry=Entry(self.ekran_gider,font=("Arial", 20))
        self.gider__ismi_entry.place(x=350,y=100)
        self.gider_lbl=Label(self.ekran_gider,text="GİDER TUTARI GİRİNİZ: ", font=("Arial", 20), bg="#666666", fg="gold")
        self.gider_lbl.place(x=5,y=200)
        self.gider_tutar_entry=Entry(self.ekran_gider,font=("Arial", 20))
        self.gider_tutar_entry.place(x=350,y=200)
        self.gider_ekle_btn=Button(self.ekran_gider, text="GİDER EKLE",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.gider_ekle)
        self.gider_ekle_btn.place(x=5,y=300)
        self.gider_sorgu_btn=Button(self.ekran_gider, text="GİDERLERİ GÖSTER",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.gider_sorgu)
        self.gider_sorgu_btn.place(x=200,y=300)
        self.gider_sil_btn=Button(self.ekran_gider, text="GİDER SİL",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.gider_sil)
        self.gider_sil_btn.place(x=500,y=300)
        self.temizle_gider=Button(self.ekran_gider, text="TEMİZLE",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.gdr_temizle)
        self.temizle_gider.place(x=680,y=300)
        self.geri_don_btn=Button(self.ekran_gider, text="GERİ DÖN",font=("Arial",20),background="#9b9b9b",fg="gold",command=self.gdr_geri_don)
        self.geri_don_btn.place(x=850,y=300)
        self.islem_gider_ayar_ekran_btn=Button(self.ekran_gider, text="İŞLEM FİYATLARINI DEĞİŞTİR",font=("Arial",14),background="#9b9b9b",fg="gold",command=self.islem_gider_degistir)
        self.islem_gider_ayar_ekran_btn.place(x=680,y=200)
        self.listbox_frame = None
        
        

        




    def gider_ekle(self):
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)
        self.sayfa=self.excel_dosya["GELİR"]
        self.brüt_sayfa=self.excel_dosya["NET GELİR"]
        self.sayfa_gider=self.excel_dosya["GİDER"]
        self.gider_ismi_entry_get = self.gider__ismi_entry.get()
        self.gider_tutar_entry_get = self.gider_tutar_entry.get()
        if self.gider_ismi_entry_get is None or self.gider_ismi_entry_get == "":
            messagebox.showerror("HATA", "GİDER İSMİ GİRMEDİNİZ")
        elif self.gider_tutar_entry_get is None or self.gider_tutar_entry_get == "":
            messagebox.showerror("HATA", "GİDER TUTARI GİRMEDİNİZ")
        try: 
            gider_tutar_int = int(self.gider_tutar_entry_get)
            self.max_satir = self.sayfa_gider.max_row
            self.gider_satir_sayisi = self.max_satir + 1  
            for satir in range(1, self.max_satir + 1):
                if all(self.sayfa_gider.cell(row=satir, column=col).value is None for col in range(1, 8)):
                    self.gider_satir_sayisi = satir  
                    break  
            self.gunun_tarihi = datetime.now().strftime("%d.%m.%Y")
            self.sayfa_gider.cell(row=self.gider_satir_sayisi, column=1, value=self.gunun_tarihi)
            self.sayfa_gider.cell(row=self.gider_satir_sayisi, column=2, value=self.gider_ismi_entry_get)
            self.sayfa_gider.cell(row=self.gider_satir_sayisi, column=3, value=gider_tutar_int)
            self.excel_dosya.save(self.excel_dosya_yolu)
            self.gdr_eklendi=Label(self.ekran_gider,text="GİDER BAŞARIYLA EKLENDİ ", font=("Arial", 20), bg="#666666", fg="gold")
            self.gdr_eklendi.place(x=100,y=400)
        except ValueError:
            messagebox.showerror("HATA", "GİDER TUTARI TAM SAYI OLMALIDIR")
                    
    def islem_gider_degistir(self):
        if hasattr(self, 'ekran_gider') and self.ekran_gider:
            try:
                if self.ekran_gider.winfo_exists():
                    self.ekran_gider.destroy()
            except tk.TclError:
                pass 
        
        self.islem_gider_ayar_ekran=Tk()
        self.islem_gider_ayar_ekran.title("Klinik Adı")
        self.islem_gider_ayar_ekran.configure(bg="#666666")
        self.islem_gider_ayar_ekran.geometry("1000x650+500+200")
        self.icon_yolu = self.get_resource_path("Klinik Logo resmi")
        
        try:
            # İkonu yükle
            self.ekran1.iconbitmap(self.icon_yolu)
        except Exception as e:
            print(f"Error loading icon: {e}")
        self.islem_degistiri_onceki_fiyat=Label(self.islem_gider_ayar_ekran,text="İŞLEM: ", font=("Arial", 18), bg="#666666", fg="white")
        self.islem_degistiri_onceki_fiyat.place(x=300,y=75)
        self.islem_degistiri_onceki_fiyat__isim_lbl=Label(self.islem_gider_ayar_ekran,text="", font=("Arial", 18), bg="#666666", fg="white")
        self.islem_degistiri_onceki_fiyat__isim_lbl.place(x=450,y=75)
        self.fiyat=Label(self.islem_gider_ayar_ekran,text="ESKİ FİYAT: ", font=("Arial", 18), bg="#666666", fg="white")
        self.fiyat.place(x=450,y=150)
        self.islem_degistiri_onceki_fiyat_lbl=Label(self.islem_gider_ayar_ekran,text="", font=("Arial", 18), bg="#666666", fg="white")
        self.islem_degistiri_onceki_fiyat_lbl.place(x=650,y=150)
        self.islem_degistilecek_lbl=Label(self.islem_gider_ayar_ekran,text="YENİ FİYAT: ", font=("Arial", 20), bg="#666666", fg="white")
        self.islem_degistilecek_lbl.place(x=450,y=300)
        self.islem_degistilecek_entry=Entry(self.islem_gider_ayar_ekran,font=("Arial", 20))

        self.var = tk.StringVar()  
        self.var.trace("w", self.format_entry_value)
        self.islem_degistilecek_entry.config(textvariable=self.var)  

        self.islem_degistilecek_entry.place(x=620,y=300)
        self.fiyat_degistir=Button(self.islem_gider_ayar_ekran, text="FİYAT EKLE",font=("Arial",20),background="#9b9b9b",fg="white",command=self.islem_degistir_fonk)
        self.fiyat_degistir.place(x=300,y=400)
        self.fiyat_degisiklik_kontrol=Label(self.islem_gider_ayar_ekran,text="", font=("Arial", 20), bg="#666666", fg="white")
        self.fiyat_degisiklik_kontrol.place(x=350,y=500)
        self.islem_gider_ayar_geridön_btn=Button(self.islem_gider_ayar_ekran, text="GERİ DÖN",font=("Arial",20),background="#9b9b9b",fg="white",command=self.islem_gider_ayar_geridön)
        self.islem_gider_ayar_geridön_btn.place(x=500,y=400)
        self.alt_kategoriler_gider_ayar = {
            "GÜLÜŞ TASARIM": ["ZİRKONYUM", "E-MAX", "LAMİNATE(YAPRAK PORSELEN)", "MONOLİTİK ZİRKONYUM","PORSELEN(METAL DESTEKLİ PROTEZ)","Tİ BASE MULTİ UNİT PORSELEN","Tİ BASE MULTİ ÜNİT MONOLİTİK ZİRKON","Tİ-BASE","Ti-BAR"],
            "PROTEZ": ["TEK ÇENE TOTAL PROTEZ", "TEK ÇENE HAREKETLİ İSKELET PROTEZ", "TEK ÇENE HİBRİT PROTEZ+BAR","TEK ÇENE HİBRİT PROTEZ+TİTANYUM BAR","OVERDENTURE PROTEZ","GECE PLAĞI","MULTİ UNİT"],
            "RESTO": ["ESTETİK DOLGU", "BONDİNG", "DOLGU","KUAFAJ","DİESTEMA KAPAMA","HASSASİYET GİDERİCİ","MASSETER BOTOKS+GECE PLAĞI","KANAL TEDAVİSİ+DOLGU","LAZER BLEACHİNG+DİŞ TAŞI TEMİZLİĞİ","KANAL YENİLEME"],
            "ÇOCUK": ["KANAL", "YER TUTUCU", "SÜT DİŞ ÇEKİMİ","DOLGU KOMPOMER","AMPUTASYON"],
            "CERRAHİ": ["DİŞ ÇEKİM","KOMPLİKE DİŞ ÇEKİM","SÜRMÜŞ YİRMİLİK DİŞ ÇEKİM","YARI GÖMÜLÜ DİŞ ÇEKİMİ","TAM GÖMÜLÜ DİŞ ÇEKİM","DİSTOANGULER GÖMÜLÜ DİŞ ÇEKİM","HORİZONTAL GÖMÜLÜ DİŞ ÇEKİM","KİSTEKTOMİ(KİST TEMİZLİĞİ)"],
            "PERİDONTOLOJİ": ["DETPOL", "TEK SEANS KÜRETAJ","ÇİFT SEANS KÜRETAJ","UZMAN KÜRETAJ","FRENEKTOMİ(UZMAN)","FLAP KÜRETAJ(UZMAN)","SDG(SERBEST DİŞETİ GREFT DİŞ BAŞI)"],
            "İMPLANT": ["BİLİM İMPLANT(YERLİ)","MEDİGMA(ALMAN)","OSSTEM(GÜNEY KORE)","MEDENTİKA(ALMAN)","STRAUMANN(İSVİÇRE)","AÇIK SİNÜS AMELİYATI(TEK TARAF)","KAPALI SİNÜS AMELİYATI","KEMİK TOZU(1cc)","MEMBRAN"]
        }
        self.secenek_durumlari_gider = {}  

        self.val_gdr = tk.StringVar()
        self.val_gdr.set("YAPILACAK İŞLEM SEÇİNİZ")
        
        # Dropdown menu
        self.islem_secenek_gdr = tk.OptionMenu(self.islem_gider_ayar_ekran, self.val_gdr, *self.alt_kategoriler_gider_ayar.keys(), command=self.gdr_ayar)
        self.islem_secenek_gdr.place(x=5, y=75)
        self.islem_secenek_gdr.configure(font=("Arial", 13), bg="#666666", fg="white")
    def format_entry_value(self,*args):
        value = self.islem_degistilecek_entry.get().replace('.', '')  
        try:
            # Sayıyı formatla
            value = f"{int(value):,}".replace(",", ".") 
        except ValueError:
            value = ""  

       
        self.islem_degistilecek_entry.delete(0, tk.END)
        self.islem_degistilecek_entry.insert(0, value)

    def gdr_ayar(self, deger):
        self.islem_fiyatları_txt=self.get_resource_path(r"klinik islem ücretleri.txt")
        # Önceki radiobutton'ları temizle
        for widget in self.islem_gider_ayar_ekran.winfo_children():
            if isinstance(widget, tk.Radiobutton):
                widget.destroy()

        # Seçilen kategoriye ait alt kategorileri al
        alt_kategoriler = self.alt_kategoriler_gider_ayar.get(deger, [])
        self.secenek_durumlari_gider[deger] = tk.StringVar()  # Yeni kategori için StringVar
        y_offset = 120  # İlk radiobutton'ın y-koordinatı
        for x, alt_kategori in enumerate(alt_kategoriler):
            radio_button = tk.Radiobutton(
                self.islem_gider_ayar_ekran,
                text=alt_kategori,
                variable=self.secenek_durumlari_gider[deger],  # Tüm seçenekleri tek bir StringVar ile kontrol et
                value=alt_kategori,  # Seçilen değeri alt kategori olarak ayarla
                font=("Arial", 13, "bold"),
                bg="#666666",
                fg="white",
                selectcolor="black",
                activebackground="#cccccc",
                activeforeground="blue",
                command=lambda var=self.secenek_durumlari_gider[deger]: self.radio_button_secilen(var)
            )
            radio_button.place(x=6, y=y_offset + x * 30)  
    def radio_button_secilen(self, selected_var):
        self.islem_fiyatları_txt=self.get_resource_path(r"klinik islem ücretleri.txt")
        self.selected_value = selected_var.get()
        with open(self.islem_fiyatları_txt, "r", encoding="utf-8") as dosya:
            for satir in dosya:
                sutunlar = satir.strip().split("=")
                if self.selected_value in sutunlar:
                    self.islem_degistiri_onceki_fiyat__isim_lbl.configure(text=self.selected_value)
                    deger=sutunlar[1]
                    dondur=deger[::-1]
                    numaraeklenmis = [dondur[i:i+3] for i in range(0, len(dondur), 3)]
                    gruplaribirlestir= '.'.join(numaraeklenmis)[::-1]
                    self.islem_degistiri_onceki_fiyat_lbl.configure(text=f"{gruplaribirlestir} TL")


    def islem_degistir_fonk(self):
        self.islem_fiyatları_txt=self.get_resource_path(r"klinik islem ücretleri.txt")
        try:
            self.yeni_fiyat_get = self.islem_degistilecek_entry.get()
            self.yeni_fiyat_get = int(self.yeni_fiyat_get.replace('.', ''))
            
            with open(self.islem_fiyatları_txt, "r", encoding="utf-8") as dosya:
                satirlar = dosya.readlines()

            # Satırlarda değişiklik yap
            with open(self.islem_fiyatları_txt, "w", encoding="utf-8") as dosya:
                for satir in satirlar:
                    sutunlar = satir.strip().split("=")
                    
                    # Eğer selected_value mevcutsa, ikinci sütunu değiştir
                    if self.selected_value in sutunlar:
                        sutunlar[1] = str(self.yeni_fiyat_get)  # int'i str'ye çevir
                        satir = "=".join(sutunlar) + "\n"
                    
                    # Satırı dosyaya yaz
                    dosya.write(satir)
            
            self.fiyat_degisiklik_kontrol.configure(text="FİYAT BAŞARIYLA DEĞİŞTİRİLDİ")
            
        except ValueError:
            messagebox.showerror("HATA", "İŞLEM SEÇMEDİNİZ VEYA DEĞER GİRMEDİNİZ")
    def islem_gider_ayar_geridön(self):
        if hasattr(self, 'islem_gider_ayar_ekran') and self.islem_gider_ayar_ekran:  
            try:  
                if self.islem_gider_ayar_ekran.winfo_exists():  
                    self.islem_gider_ayar_ekran.destroy()    
            except tk.TclError:  
                pass 
        self.gider_sayfasi()
    
    def gider_sorgu(self):
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)
        self.sayfa=self.excel_dosya["GELİR"]
        self.brüt_sayfa=self.excel_dosya["NET GELİR"]
        self.sayfa_gider=self.excel_dosya["GİDER"]
        self.gider_ismi_entry_get=self.gider__ismi_entry.get().upper()
        # Gider sorgu işlemi
        aylar = [
                                "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                                "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"
                        ]
        

        self.toplam_gdr = 0  # Toplam gideri sıfırla
        self.gider_tree = ttk.Treeview(self.ekran_gider, columns=("Tarih", "İsim", "İşlem Tutarı",   
                    "Doktor1", "Doktor2",   
                    "Doktor3", "Doktor4",   
                    "Doktor5"),
            show="headings")  

        # Başlık ayarlarını yap  
        for col in self.gider_tree["columns"]:  
            self.gider_tree.heading(col, text=col)  # Her bir sütun başlığını ayarla  
        for col in self.gider_tree["columns"]:  
            self.gider_tree.column(col, anchor='center')
       
        
        self.gider_tree.place(x=5,y=400)
        self.satir_sayisi=1
        baslangıc_sutün=2
        bitis_sutün=0
        while True:  
            # Hücrede bir değer olup olmadığını kontrol et  
            self.veri_varmi = self.sayfa_gider.cell(row=self.satir_sayisi, column=1).value is not None  

            # 1 ile 7 arasındaki sütunlar için döngü  
            for col in range(1, 8):   
                self.ay_deger = self.sayfa_gider.cell(row=self.satir_sayisi, column=col).value  
                if self.ay_deger in aylar:  
                    baslangıc_sutün = self.satir_sayisi + 1
                  

            # Veri yoksa döngüyü bitir  
            if not self.veri_varmi:  
                bitis_sutün = self.satir_sayisi  
                break  
                
            # Satır sayısını artır  
            self.satir_sayisi += 1
       
        
        for satir in range(baslangıc_sutün, bitis_sutün):  
            entry_values = []  # Değerleri toplamak için bir liste  
            for i in range(1, 9):  
                deger = self.sayfa_gider.cell(row=satir, column=i).value  
                # Boş değerleri kontrol et, boşsa '' ekle  
                if deger is None:  # Eğer deger None ise  
                    entry_values.append('')  # Boş değer için içeriğe '' ekle  
                else:  
                    entry_values.append(deger)  # Değer mevcutsa ekle  

            if entry_values:  # entry_values boş değilse  
                self.gider_tree.insert("", tk.END, values=entry_values)  # Dolu olan değerleri ekle   
       

    def gider_sil(self):
        # Excel dosyasını yükle
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya = openpyxl.load_workbook(self.excel_dosya_yolu)
        self.sayfa_gider = self.excel_dosya["GİDER"]
        aylar = [
                                "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                                "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"
                        ]

        # Treeview'da seçili öğeyi al
        secili_index = self.gider_tree.selection()
        secili_veri = self.gider_tree.item(secili_index, "values")
        secili_veri_liste = list(secili_veri)
        if not secili_index:
            messagebox.showerror("HATA", "LÜTFEN SİLİNECEK BİR DEĞER SEÇİNİZ")
            return

        # Seçili öğenin verisini al
        secili_veri = self.gider_tree.item(secili_index, "values")
        if not secili_veri:
            messagebox.showerror("HATA", "SİLİNECEK DEĞERİN VERİSİ ALGILANMADI")
            return
        baslangıc_sutün=2
        bitis_sutün=0
        while True:  
            # Hücrede bir değer olup olmadığını kontrol et  
            self.veri_varmi = self.sayfa_gider.cell(row=self.satir_sayisi, column=1).value is not None  

            # 1 ile 7 arasındaki sütunlar için döngü  
            for col in range(1, 8):   
                self.ay_deger = self.sayfa_gider.cell(row=self.satir_sayisi, column=col).value  
                if self.ay_deger in aylar:  
                    baslangıc_sutün = self.satir_sayisi + 1  

            # Veri yoksa döngüyü bitir  
            if not self.veri_varmi:  
                bitis_sutün = self.satir_sayisi  
                break  
                
            # Satır sayısını artır  
            self.satir_sayisi += 1


        for satir in range(baslangıc_sutün, bitis_sutün):
            deger=self.sayfa_gider.cell(row=satir,column=2).value
            if deger==secili_veri_liste[1]:
                self.sayfa_gider.delete_rows(satir)
        
        # Excel dosyasını kaydet
        self.excel_dosya.save(self.excel_dosya_yolu)
        messagebox.showinfo("BAŞARILI", "GİDER BAŞARIYLA SİLİNDİ")

        # Gider sorgusunu yenile
        self.gider_sorgu()
    def gdr_temizle(self):
            if hasattr(self, 'gider_tree') and self.gider_tree:
                self.gider_tree.destroy()
            self.gider__ismi_entry.delete(0, END)
            self.gider_tutar_entry.delete(0, END)

    def gdr_geri_don(self):
        if hasattr(self, 'ekran_gider') and self.ekran_gider:
            try:
                if self.ekran_gider.winfo_exists():
                    self.ekran_gider.destroy()
            except tk.TclError:
                pass 
        self.islemler()

    def bütün_excel_goster(self):
        self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
        self.excel_dosya = openpyxl.load_workbook(self.excel_dosya_yolu)
        messagebox.showinfo("BEKLEYİN","EXCEL DOSYASI AÇILIYOR LÜTFEN BEKLEYİN")
        try:
            subprocess.Popen(["start", "excel", self.excel_dosya_yolu], shell=True)
        except Exception as e:
           messagebox.showerror("HATA","EXCEL DOSYASI BİR HATADAN DOLAYI AÇILAMIYOR")
    
    def aylikciro(self):
                self.excel_dosya_yolu = self.get_resource_path(r"excel tablo yolu ")
                self.excel_dosya=openpyxl.load_workbook(self.excel_dosya_yolu)
                self.sayfa=self.excel_dosya["GELİR"]
                self.brüt_sayfa=self.excel_dosya["NET GELİR"]
                self.sayfa_gider=self.excel_dosya["GİDER"]
                self.aylik_ciro_ekran=Tk()
                self.aylik_ciro_ekran.title("Klinik Adı")
                self.aylik_ciro_ekran.configure(bg="#666666")
                self.aylik_ciro_ekran.geometry("2200x750+20+50")
                self.icon_yolu = self.get_resource_path("Klinik Logo resmi")
        
                try:
                    # İkonu yükle
                    self.ekran1.iconbitmap(self.icon_yolu)
                except Exception as e:
                    print(f"Error loading icon: {e}")

                aylar = [
                                "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                                "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"
                        ]

                self.ay = self.bugün.month 
                ay = aylar[self.ay - 1]  
                self.max_satir_glr=self.sayfa_gider.max_row
                self.satir_sayisi=2267
                self.toplam_islem_ücreti=0
                self.toplam_ödenen=0
                self.aylik_net_gelir=0
                self.aylik_maliyet=0
                self.doktor1_prim=0
                self.doktor2_prim=0
                self.doktor3_prim=0
                self.doktor4_prim=0
                self.doktor5_prim=0
                while True:
                    # Satırda veri olup olmadığını kontrol et
                    self.veri_varmi = any(self.sayfa.cell(row=self.satir_sayisi, column=col).value is not None for col in range(1, 16))
                    
                    # Eğer satırda ay değeri varsa, başlangıç sütununu belirle
                    for col in range(1, 16):
                        self.ay_deger = self.sayfa.cell(row=self.satir_sayisi, column=col).value
                        if self.ay_deger in aylar:
                            baslangic_sutun = self.satir_sayisi + 1
                    
                    # Eğer satırda veri yoksa, bitiş sütununu belirle ve döngüden çık
                    if not self.veri_varmi:
                        bitis_sutun = self.satir_sayisi
                        break
                    
                    # Bir sonraki satıra geç
                    self.satir_sayisi += 1

                # Benzersiz kayıtları saklamak için bir set oluştur
                benzersiz_kayitlar = set()

                # Sonuç listesi
                self.dizi = []

                # Başlangıç sütunundan bitiş sütununa kadar döngü
                i = baslangic_sutun
                while i < bitis_sutun:
                    gecici_dizi = []
                    
                    isim1 = self.sayfa.cell(row=i, column=3).value
                    islem1 = self.sayfa.cell(row=i, column=4).value
                    gider1=self.sayfa.cell(row=i, column=5).value
                    islem_tutar1 = self.sayfa.cell(row=i, column=12).value
                    
                    if islem_tutar1 is not None:
                        islem_tutar1 = int(islem_tutar1.replace("₺", "").replace(".", ""))
                    else:
                        islem_tutar1 = 0  
                    if gider1 is not None:
                        gider1 = int(gider1)
                    else:
                        gider1=0
                
                    kayit = (isim1, islem1, islem_tutar1,gider1)
                    if kayit not in benzersiz_kayitlar:
                        benzersiz_kayitlar.add(kayit)
                        gecici_dizi.extend([isim1, islem1, islem_tutar1,gider1])
                        self.dizi.append(gecici_dizi)
                    
                    i += 1
                
               
                
                for kayıt in self.dizi:
                    deger=kayıt[2]
                    deger = int(deger)
                    self.toplam_islem_ücreti+=deger

                    
                for satir in range(baslangic_sutun,bitis_sutun + 1): 
                    deger = self.sayfa.cell(row=satir, column=15).value  
                    if deger in aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.aylik_net_gelir += temiz_deger 
                    else:
                        if deger is not None:
                            self.aylik_net_gelir += int(deger)  
                for kayıt in self.dizi:
                    deger=kayıt[3]
                    deger = int(deger)
                    self.aylik_maliyet+=deger        
                 

                for satir in range(baslangic_sutun,bitis_sutun + 1): 
                    deger = self.sayfa.cell(row=satir, column=7).value  
                    if deger in aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor1_prim += temiz_deger 
                    else:
                        if deger is not None:
                            self.doktor1_prim += int(deger) 

                for satir in range(baslangic_sutun,bitis_sutun + 1): 
                    deger = self.sayfa.cell(row=satir, column=8).value  
                    if deger in aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor2_prim += temiz_deger 
                    else:
                        if deger is not None:
                            self.doktor2_prim += int(deger)  

                for satir in range(baslangic_sutun,bitis_sutun + 1): 
                    deger = self.sayfa.cell(row=satir, column=9).value  
                    if deger in aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor3_prim += temiz_deger 
                    else:
                        if deger is not None:
                            self.doktor3_prim += int(deger)  

                for satir in range(baslangic_sutun,bitis_sutun + 1): 
                    deger = self.sayfa.cell(row=satir, column=10).value  
                    if deger in aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor4_prim += temiz_deger 
                    else:
                        if deger is not None:
                            self.doktor4_prim += int(deger)  

                for satir in range(baslangic_sutun,bitis_sutun + 1): 
                    deger = self.sayfa.cell(row=satir, column=11).value  
                    if deger in aylar:  
                        continue
                    elif isinstance(deger, str) and "₺" in deger: 
                        temiz_deger = int(deger.replace("₺", "").strip()) 
                        self.doktor5_prim += temiz_deger
                    else:
                        if deger is not None:
                            self.doktor5_prim += int(deger) 
                        
                self.toplam_ay_islem_noktali=f"{self.toplam_islem_ücreti:,}".replace(",", ".")
                self.aylik_net_gelir_noktali=f"{self.aylik_net_gelir:,}".replace(",", ".")
                self.aylik_maliyet_noktali=f"{self.aylik_maliyet:,}".replace(",", ".")
                self.doktor1_prim_noktali=f"{self.doktor1_prim:,}".replace(",", ".")
                self.doktor2_prim_noktali=f"{self.doktor2_prim:,}".replace(",", ".")
                self.doktor3_prim_noktali=f"{self.doktor3_prim:,}".replace(",", ".")
                self.doktor4_prim_noktali=f"{self.doktor4_prim:,}".replace(",", ".")
                self.doktor5_prim_noktali=f"{self.doktor5_prim:,}".replace(",", ".")


                lbl_ay = Label(self.aylik_ciro_ekran, text="AY ||", font=("Arial", 15), bg="#666666", fg="gold")
                lbl_ay.grid(row=0, column=0, padx=10)

                lbl_ay_grd=Label(self.aylik_ciro_ekran, text=ay, font=("Arial", 15), bg="#666666", fg="gold")
                lbl_ay_grd.grid(row=1, column=0, padx=10)

                lbl_toplam_ay_islem = Label(self.aylik_ciro_ekran, text="TOPLAM İŞLEM ÜCRETİ ||", font=("Arial", 15), bg="#666666", fg="gold")
                lbl_toplam_ay_islem.grid(row=0, column=1, padx=10)

                lbl_toplam_ay_islem_grd=Label(self.aylik_ciro_ekran, text=self.toplam_ay_islem_noktali, font=("Arial", 15), bg="#666666", fg="gold")
                lbl_toplam_ay_islem_grd.grid(row=1, column=1, padx=10)

                lbl_aylik_maliyet = Label(self.aylik_ciro_ekran, text="MALİYET ||", font=("Arial", 15), bg="#666666", fg="gold")
                lbl_aylik_maliyet.grid(row=0, column=2, padx=10)

                lbl_aylik_maliyet_grd=Label(self.aylik_ciro_ekran, text=self.aylik_maliyet_noktali, font=("Arial", 15), bg="#666666", fg="gold")
                lbl_aylik_maliyet_grd.grid(row=1, column=2, padx=10)

                lbl_aylik_net_gelir = Label(self.aylik_ciro_ekran, text="NET GELİR ||", font=("Arial", 15), bg="#666666", fg="gold")
                lbl_aylik_net_gelir.grid(row=0, column=3, padx=10)

                lbl_aylik_net_gelir_grd=Label(self.aylik_ciro_ekran, text=self.aylik_net_gelir_noktali, font=("Arial", 15), bg="#666666", fg="gold")
                lbl_aylik_net_gelir_grd.grid(row=1, column=3, padx=10)

                lbl_doktor1_prim = Label(self.aylik_ciro_ekran, text="Doktor1 ||", font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor1_prim.grid(row=0, column=4, padx=10)

                lbl_doktor1_prim_grd=Label(self.aylik_ciro_ekran, text=self.doktor1_prim_noktali, font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor1_prim_grd.grid(row=1, column=4, padx=10)

                lbl_doktor2_prim = Label(self.aylik_ciro_ekran, text="Doktor2 ||", font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor2_prim.grid(row=0, column=5, padx=10)

                lbl_doktor2_prim_grd=Label(self.aylik_ciro_ekran, text=self.doktor2_prim_noktali, font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor2_prim_grd.grid(row=1, column=5, padx=10)

                lbl_doktor3_prim = Label(self.aylik_ciro_ekran, text="Doktor3 ||", font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor3_prim.grid(row=0, column=6, padx=10)

                lbl_doktor3_prim=Label(self.aylik_ciro_ekran, text=self.doktor3_prim_noktali, font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor3_prim.grid(row=1, column=6, padx=10)

                lbl_doktor4_prim = Label(self.aylik_ciro_ekran, text="Doktor4 ||", font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor4_prim.grid(row=0, column=7, padx=10)

                lbl_doktor4_prim=Label(self.aylik_ciro_ekran, text=self.doktor4_prim_noktali, font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor4_prim.grid(row=1, column=7, padx=10)

                lbl_doktor5_prim = Label(self.aylik_ciro_ekran, text="Doktor5 ||", font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor5_prim.grid(row=0, column=8, padx=10)

                lbl_doktor5_prim_grd=Label(self.aylik_ciro_ekran, text=self.doktor5_prim_noktali, font=("Arial", 15), bg="#666666", fg="gold")
                lbl_doktor5_prim_grd.grid(row=1, column=8, padx=10)

                
                
                



        
   
                
            
            

Uygulama()
